"""测评报告导出：Excel（多 sheet 明细）+ PDF（可视化汇报）。

evalscope 1.8.1 输出字段：
  accuracy: { accuracy, num, by_subject, by_category, analysis, dataset_description, perf_metrics, few_shot, repro, raw }
  performance: { sweep, best, lowest_latency, recommend, warnings, profile }
  context_scan: { sweep, concurrency, requests_per_level }
"""
import os
import time

REPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "reports")
os.makedirs(REPORT_DIR, exist_ok=True)


def _ts():
    return time.strftime("%Y%m%d_%H%M%S")


def _fmt_time(epoch):
    if not epoch:
        return "-"
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(epoch))


def _s(v):
    """安全字符串：None → '-'"""
    if v is None:
        return "-"
    return str(v)


def _n(v, default=0):
    """安全数字：None/NaN → default"""
    import math
    if v is None:
        return default
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return default
        return v
    except (ValueError, TypeError):
        return default


def _dur(sec):
    """秒 → 可读时长"""
    if sec is None:
        return "-"
    sec = int(sec)
    if sec < 60:
        return f"{sec}s"
    if sec < 3600:
        return f"{sec // 60}m{sec % 60}s"
    return f"{sec // 3600}h{(sec % 3600) // 60}m"


# ============ Excel 导出 ============

def export_excel(detail: dict) -> str:
    """把任务结果导出为 Excel 多 sheet 报告。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    name = detail.get("name", "report")
    result = detail.get("result", {})
    config = detail.get("config", {})

    # ── 配色 ──
    PRIMARY = "4F46E5"
    PRIMARY_LT = "EEF0FB"
    ACCENT = "0E7490"
    DARK = "1F2433"
    DIM = "6B7280"
    LINE = "E5E7EB"
    WHITE = "FFFFFF"
    OK_GREEN = "D1FAE5"
    WARN_AMBER = "FEF3C7"
    ROW_ALT = "F9FAFB"

    hdr_fill = PatternFill("solid", fgColor=PRIMARY)
    hdr_font = Font(bold=True, color=WHITE, name="Arial", size=10)
    lt_fill = PatternFill("solid", fgColor=PRIMARY_LT)
    dim_font = Font(color=DIM, name="Arial", size=10)
    bold_font = Font(bold=True, name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    mono_font = Font(name="Consolas", size=10)
    kpi_font = Font(bold=True, color=PRIMARY, name="Arial", size=22)
    title_font = Font(bold=True, color=WHITE, name="Arial", size=18)
    h2_font = Font(bold=True, color=PRIMARY, name="Arial", size=12)

    thin = Side(style="thin", color=LINE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border

    def style_body(ws, start_row, end_row, ncols):
        for r in range(start_row, end_row + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.font = body_font
                if (r - start_row) % 2:
                    cell.fill = PatternFill("solid", fgColor=ROW_ALT)

    def auto_width(ws, ncols, min_w=10, max_w=55):
        for c in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(c)].width = min_w

    # ===== Sheet 1: 概览 =====
    ws = wb.active
    ws.title = "概览"
    ws.sheet_properties.tabColor = PRIMARY

    # 标题横幅
    ws.merge_cells("A1:E1")
    ws["A1"] = "大模型测评报告"
    ws["A1"].font = title_font
    ws["A1"].fill = PatternFill("solid", fgColor=PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    ws.append([])

    # 任务信息表
    acc = result.get("accuracy", {})
    perf = result.get("performance", {})
    ctx = result.get("context_scan", {})
    pc = config.get("perf", {}) or {}

    info = [
        ("任务名称", name),
        ("模型", config.get("model", "-")),
        ("接口地址", (config.get("base_url") or "-")[:80]),
        ("接口格式", config.get("api_format", "-")),
        ("关闭思考", "是" if config.get("disable_thinking") else "否"),
        ("状态", detail.get("status", "-")),
        ("运行时长", _dur(detail.get("duration"))),
        ("创建时间", _fmt_time(detail.get("created"))),
        ("完成时间", _fmt_time(detail.get("finished_at"))),
    ]
    for i, (k, v) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=k).font = bold_font
        ws.cell(row=i, column=1).fill = lt_fill
        ws.cell(row=i, column=1).border = border
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        ws.cell(row=i, column=2, value=str(v)).font = body_font
        ws.cell(row=i, column=2).border = border
        for c in range(3, 5):
            ws.cell(row=i, column=c).border = border
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50

    # 测试配置摘要
    row = len(info) + 4
    ws.cell(row=row, column=1, value="评测配置").font = h2_font
    row += 1
    cfg_items = [
        ("精度数据集", ", ".join(config.get("accuracy_datasets", [])[:12]) or "无"),
        ("学科筛选", _fmt_subjects(config.get("dataset_subjects", {}))),
        ("Few-shot", str(config.get("few_shot", 0))),
        ("抽样上限", _s(config.get("sample_limit") or "全量")),
        ("精度 max_tokens", _s(config.get("acc_max_tokens") or "自动")),
        ("温度", str(config.get("acc_temperature", 0))),
        ("失败重试", str(config.get("max_retries", 2))),
        ("性能压测", _fmt_perf_config(config)),
        ("上下文扫描", _fmt_ctx_config(config)),
    ]
    for k, v in cfg_items:
        ws.cell(row=row, column=1, value=k).font = dim_font
        ws.cell(row=row, column=2, value=v).font = body_font
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        row += 1

    # KPI 卡片
    row += 1
    ws.cell(row=row, column=1, value="执行摘要").font = h2_font
    row += 1
    kpis = _build_kpis(result)
    for i, (val, lbl) in enumerate(kpis):
        c = i * 2 + 1
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c + 1)
        ws.cell(row=row, column=c, value=str(val)).font = kpi_font
        ws.cell(row=row, column=c).alignment = center
        ws.cell(row=row, column=c).fill = lt_fill
        ws.cell(row=row, column=c).border = border
        ws.cell(row=row + 1, column=c).value = lbl
        ws.cell(row=row + 1, column=c).font = dim_font
        ws.cell(row=row + 1, column=c).alignment = center
        ws.cell(row=row + 1, column=c).fill = lt_fill
        ws.cell(row=row + 1, column=c).border = border
        ws.merge_cells(start_row=row + 1, start_column=c, end_row=row + 1, end_column=c + 1)
        ws.cell(row=row, column=c + 1).fill = lt_fill
        ws.cell(row=row, column=c + 1).border = border
        ws.cell(row=row + 1, column=c + 1).fill = lt_fill
        ws.cell(row=row + 1, column=c + 1).border = border
    for c in range(1, len(kpis) * 2 + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ===== Sheet 2: 精度总览 =====
    if acc:
        ws2 = wb.create_sheet("精度总览")
        ws2.sheet_properties.tabColor = "10B981"
        acc_rows = [["数据集", "准确率", "题目数", "学科数", "类别数", "Few-shot", "后端"]]
        for ds, d in acc.items():
            shot = _n(d.get("few_shot", 0))
            shot_label = f"{shot}-shot" if shot else "0-shot"
            acc_rows.append([
                ds,
                f"{d['accuracy']}%" if d.get("accuracy") is not None else "见详情",
                _n(d.get("num") or d.get("total")),
                len(d.get("by_subject", {}) or {}),
                len(d.get("by_category", {}) or {}),
                shot_label,
                d.get("backend", "evalscope"),
            ])
        for i, r in enumerate(acc_rows):
            for j, v in enumerate(r):
                ws2.cell(row=i + 1, column=j + 1, value=v)
            if i == 0:
                style_header(ws2, 1, len(r))
            else:
                style_body(ws2, 2, len(acc_rows), len(r))
        ws2.freeze_panes = "A2"
        for c in range(1, 8):
            ws2.column_dimensions[get_column_letter(c)].width = [16, 12, 10, 8, 8, 10, 14][c - 1]

    # ===== Sheet 3: 分学科/类别明细 =====
    has_subj = any((d.get("by_subject") or d.get("by_category")) for d in acc.values())
    if has_subj:
        ws3 = wb.create_sheet("分科明细")
        ws3.sheet_properties.tabColor = "8B5CF6"
        headers = ["数据集", "类别", "类别准确率", "学科", "学科准确率"]
        ws3.append(headers)
        style_header(ws3, 1, len(headers))
        row = 2
        for ds, d in acc.items():
            cats = d.get("by_category", {}) or {}
            if cats:
                for cn, cv in cats.items():
                    for sn, sv in (cv.get("subsets", {}) or {}).items():
                        cat_pct = f"{cv['score']}%" if cv.get("score") is not None else "-"
                        sub_pct = f"{sv['score']}%" if sv.get("score") is not None else "-"
                        ws3.append([ds, cn, cat_pct, sn, sub_pct])
                        row += 1
            else:
                subjs = d.get("by_subject", {}) or {}
                for sn, sv in subjs.items():
                    ws3.append([ds, "-", "-", sn, f"{sv}%"])
                    row += 1
        style_body(ws3, 2, row - 1, len(headers))
        ws3.freeze_panes = "A2"
        for c, w in enumerate([16, 16, 12, 28, 12], 1):
            ws3.column_dimensions[get_column_letter(c)].width = w

    # ===== Sheet 4: 性能明细 =====
    sweep = perf.get("sweep") if isinstance(perf, dict) else None
    if sweep:
        ws4 = wb.create_sheet("性能明细")
        ws4.sheet_properties.tabColor = "F59E0B"
        cols = ["并发", "RPS", "输出tok/s", "TPOT(ms)", "TTFT(s)",
                "平均延迟(s)", "P90(s)", "P99(s)",
                "输入tok", "输出tok", "成功", "总数", "成功率%", "错误率%"]
        ws4.append(cols)
        style_header(ws4, 1, len(cols))
        best_conc = (perf.get("best") or {}).get("concurrency")
        for i, r in enumerate(sweep):
            row_data = [
                r.get("concurrency"), _n(r.get("rps")), _n(r.get("output_tps")),
                _n(r.get("tpot_avg_ms")), _n(r.get("ttft_avg")),
                _n(r.get("latency_avg")), _n(r.get("latency_p90")),
                _n(r.get("latency_p99")),
                _n(r.get("avg_in_tokens")), _n(r.get("avg_out_tokens")),
                _s(r.get("success")), _s(r.get("total")),
                _n(r.get("success_rate")), _n(r.get("error_rate")),
            ]
            ws4.append(row_data)
            row_num = i + 2
            style_body(ws4, row_num, row_num, len(cols))
            if best_conc and r.get("concurrency") == best_conc:
                for c in range(1, len(cols) + 1):
                    ws4.cell(row=row_num, column=c).fill = PatternFill("solid", fgColor=OK_GREEN)
        ws4.freeze_panes = "A2"
        for c in range(1, len(cols) + 1):
            ws4.column_dimensions[get_column_letter(c)].width = 12

        # 推荐说明
        row = len(sweep) + 3
        rec = perf.get("recommend")
        if rec:
            ws4.cell(row=row, column=1, value="推荐并发区间").font = bold_font
            ws4.cell(row=row, column=2, value=f"{rec['min']} ~ {rec['max']}").font = body_font
            row += 1
            ws4.cell(row=row, column=1, value="推荐依据").font = bold_font
            ws4.cell(row=row, column=2, value=rec.get("basis", "-")).font = body_font
            row += 1
        for w in (perf.get("warnings") or []):
            ws4.cell(row=row, column=1, value=f"⚠ {w.get('title', '')}").font = Font(bold=True, color="D97706", name="Arial", size=10)
            ws4.cell(row=row, column=2, value=w.get("message", "")).font = body_font
            row += 1

        # 压测配置
        row += 1
        ws4.cell(row=row, column=1, value="压测配置").font = h2_font
        row += 1
        profile = perf.get("profile", {}) or {}
        for k, v in [
            ("并发档位", ", ".join(str(x) for x in (profile.get("levels") or []))),
            ("每档请求数", profile.get("requests_per_level", "-")),
            ("max_tokens", profile.get("max_tokens", "-")),
            ("temperature", profile.get("temperature", "-")),
            ("模式", "流式" if profile.get("stream") else "非流式"),
        ]:
            ws4.cell(row=row, column=1, value=k).font = dim_font
            ws4.cell(row=row, column=2, value=str(v)).font = body_font
            row += 1

    # ===== Sheet 5: 上下文扫描 =====
    ctx_sweep = ctx.get("sweep") if isinstance(ctx, dict) else None
    if ctx_sweep:
        ws5 = wb.create_sheet("上下文扫描")
        ws5.sheet_properties.tabColor = "3B82F6"
        cols = ["上下文长度", "RPS", "输出tok/s", "TTFT(s)", "TPOT(ms)",
                "平均延迟(s)", "P90(s)", "P99(s)", "成功率%"]
        ws5.append(cols)
        style_header(ws5, 1, len(cols))
        for i, r in enumerate(ctx_sweep):
            row_data = [
                r.get("context_length"), _n(r.get("rps")), _n(r.get("output_tps")),
                _n(r.get("ttft_avg")), _n(r.get("tpot_avg_ms")),
                _n(r.get("latency_avg")), _n(r.get("latency_p90")),
                _n(r.get("latency_p99")), _n(r.get("success_rate")),
            ]
            ws5.append(row_data)
            style_body(ws5, i + 2, i + 2, len(cols))
        ws5.freeze_panes = "A2"
        for c in range(1, len(cols) + 1):
            ws5.column_dimensions[get_column_letter(c)].width = 14
        # 元信息
        row = len(ctx_sweep) + 3
        ws5.cell(row=row, column=1, value="固定并发").font = bold_font
        ws5.cell(row=row, column=2, value=str(ctx.get("concurrency", "-"))).font = body_font
        ws5.cell(row=row + 1, column=1, value="每档请求数").font = bold_font
        ws5.cell(row=row + 1, column=2, value=str(ctx.get("requests_per_level", "-"))).font = body_font

    # ===== Sheet 6: 推理性能统计 =====
    any_perf_metrics = any(d.get("perf_metrics") for d in acc.values())
    if any_perf_metrics:
        ws6 = wb.create_sheet("推理性能")
        ws6.sheet_properties.tabColor = "06B6D4"
        row = 1
        for ds, d in acc.items():
            pm = d.get("perf_metrics")
            if not isinstance(pm, dict):
                continue
            ws6.cell(row=row, column=1, value=f"数据集：{ds}").font = h2_font
            row += 1
            lat = pm.get("latency", {}) or {}
            tp = pm.get("throughput", {}) or {}
            usage = pm.get("usage", {}) or {}
            in_tok = usage.get("input_tokens", {}) or {}
            out_tok = usage.get("output_tokens", {}) or {}

            sub_cols = ["指标", "P25", "P50", "P75", "P90", "P99", "均值", "标准差"]
            for j, c in enumerate(sub_cols, 1):
                ws6.cell(row=row, column=j, value=c)
            style_header(ws6, row, len(sub_cols))
            row += 1

            for label, data_map in [
                ("延迟(s)", lat), ("输入Token", in_tok), ("输出Token", out_tok),
            ]:
                vals = [
                    label,
                    _n(data_map.get("25%")), _n(data_map.get("50%")), _n(data_map.get("75%")),
                    _n(data_map.get("90%")), _n(data_map.get("99%")),
                    _n(data_map.get("mean")), _n(data_map.get("std")),
                ]
                for j, v in enumerate(vals, 1):
                    ws6.cell(row=row, column=j, value=v)
                style_body(ws6, row, row, len(sub_cols))
                row += 1

            row += 1
            ws6.cell(row=row, column=1, value=f"样本数: {pm.get('n_samples', '-')}").font = dim_font
            ws6.cell(row=row, column=2, value=f"输出吞吐: {_n(tp.get('avg_output_tps'))} tok/s").font = dim_font
            ws6.cell(row=row, column=3, value=f"请求速率: {_n(tp.get('avg_req_ps'))} req/s").font = dim_font
            row += 2
        for c in range(1, 9):
            ws6.column_dimensions[get_column_letter(c)].width = 14

    # ===== Sheet 7: 错题样本 =====
    wrong_rows = []
    for ds, d in acc.items():
        for w in d.get("wrong_samples", []):
            wrong_rows.append([ds, w.get("question", ""), w.get("expected", ""),
                               w.get("got", ""), (w.get("raw", "") or "")[:300]])
    if wrong_rows:
        ws7 = wb.create_sheet("错题样本")
        ws7.sheet_properties.tabColor = "EF4444"
        cols = ["数据集", "题目", "正确答案", "模型答案", "原始输出(节选)"]
        ws7.append(cols)
        style_header(ws7, 1, len(cols))
        for i, r in enumerate(wrong_rows):
            ws7.append(r)
            style_body(ws7, i + 2, i + 2, len(cols))
        ws7.freeze_panes = "A2"
        widths = [12, 55, 10, 10, 65]
        for c, w in enumerate(widths, 1):
            ws7.column_dimensions[get_column_letter(c)].width = w

    safe_name = "".join(c for c in name if c.isalnum() or c in "-_ ")[:60] or "report"
    path = os.path.join(REPORT_DIR, f"{safe_name}_{_ts()}.xlsx")
    wb.save(path)
    return path


# ============ PDF 导出 ============

_CN_FONT_REGISTERED = [None]


def _register_cn_font(pdfmetrics, TTFont):
    if _CN_FONT_REGISTERED[0]:
        return _CN_FONT_REGISTERED[0]
    candidates = [
        ("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc", 0),
        ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", 0),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", None),
    ]
    # also try app-local
    local = os.path.join(os.path.dirname(__file__), "fonts", "wqy-zenhei.ttc")
    if os.path.exists(local):
        candidates.insert(0, (local, 0))
    for path, idx in candidates:
        if os.path.exists(path):
            try:
                if idx is not None:
                    pdfmetrics.registerFont(TTFont("CNFont", path, subfontIndex=idx))
                else:
                    pdfmetrics.registerFont(TTFont("CNFont", path))
                _CN_FONT_REGISTERED[0] = "CNFont"
                return "CNFont"
            except Exception:
                continue
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    _CN_FONT_REGISTERED[0] = "STSong-Light"
    return "STSong-Light"


def export_pdf(detail: dict) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable, KeepTogether)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    CN = _register_cn_font(pdfmetrics, TTFont)
    PRIMARY = colors.HexColor("#4F46E5")
    PRIMARY_LT = colors.HexColor("#EEF0FB")
    DATA_CLR = colors.HexColor("#0E7490")
    INK = colors.HexColor("#1F2433")
    DIM = colors.HexColor("#6B7280")
    LINE = colors.HexColor("#E5E7EB")
    OK = colors.HexColor("#059669")
    WARN = colors.HexColor("#D97706")
    DARK_BG = colors.HexColor("#1F2433")
    WHITE = colors.white

    name = detail.get("name", "测评报告")
    result = detail.get("result", {})
    config = detail.get("config", {})

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle("t", parent=styles["Title"], fontName=CN, fontSize=22,
                             textColor=WHITE, leading=28, spaceAfter=2, alignment=TA_CENTER)
    st_sub = ParagraphStyle("s", parent=styles["Normal"], fontName=CN, fontSize=10,
                           textColor=colors.HexColor("#C7D2FE"), alignment=TA_CENTER, leading=14)
    st_h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName=CN, fontSize=13,
                          textColor=PRIMARY, spaceBefore=14, spaceAfter=6, leading=18)
    st_h3 = ParagraphStyle("h3", parent=styles["Heading3"], fontName=CN, fontSize=11,
                          textColor=INK, spaceBefore=8, spaceAfter=4, leading=14)
    st_body = ParagraphStyle("b", parent=styles["Normal"], fontName=CN, fontSize=9.5,
                            textColor=INK, leading=15)
    st_small = ParagraphStyle("sm", parent=styles["Normal"], fontName=CN, fontSize=8,
                             textColor=DIM, leading=12)
    st_kpi = ParagraphStyle("kpi", parent=styles["Normal"], fontName=CN, fontSize=20,
                           textColor=PRIMARY, alignment=TA_CENTER, leading=24)
    st_kpi_lbl = ParagraphStyle("kpil", parent=styles["Normal"], fontName=CN, fontSize=8,
                               textColor=DIM, alignment=TA_CENTER, leading=11)
    st_analysis = ParagraphStyle("an", parent=styles["Normal"], fontName=CN, fontSize=9,
                                textColor=INK, leading=15, spaceBefore=4, spaceAfter=4)

    safe_name = "".join(c for c in name if c.isalnum() or c in "-_ ")[:60] or "report"
    path = os.path.join(REPORT_DIR, f"{safe_name}_{_ts()}.pdf")

    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(PRIMARY)
        canvas.rect(0, A4[1] - 5 * mm, A4[0], 5 * mm, fill=1, stroke=0)
        canvas.setFont(CN, 7.5)
        canvas.setFillColor(DIM)
        canvas.drawString(16 * mm, 9 * mm, "LLM Eval Bench · 自动生成")
        canvas.drawRightString(A4[0] - 16 * mm, 9 * mm, f"第 {doc.page} 页")
        canvas.setStrokeColor(LINE)
        canvas.line(16 * mm, 12 * mm, A4[0] - 16 * mm, 12 * mm)
        canvas.restoreState()

    def on_first_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(DARK_BG)
        canvas.rect(0, A4[1] - 50 * mm, A4[0], 50 * mm, fill=1, stroke=0)
        canvas.setFillColor(PRIMARY)
        canvas.rect(0, A4[1] - 50 * mm, A4[0], 3 * mm, fill=1, stroke=0)
        canvas.setFont(CN, 7.5)
        canvas.setFillColor(DIM)
        canvas.drawString(16 * mm, 9 * mm, "LLM Eval Bench · 自动生成")
        canvas.drawRightString(A4[0] - 16 * mm, 9 * mm, f"第 {doc.page} 页")
        canvas.setStrokeColor(LINE)
        canvas.line(16 * mm, 12 * mm, A4[0] - 16 * mm, 12 * mm)
        canvas.restoreState()

    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
                           leftMargin=16 * mm, rightMargin=16 * mm)
    story = []
    acc = result.get("accuracy", {})
    perf = result.get("performance", {})
    ctx = result.get("context_scan", {})
    sweep = perf.get("sweep") if isinstance(perf, dict) else None
    ctx_sweep = ctx.get("sweep") if isinstance(ctx, dict) else None

    # ===== 封面标题 =====
    story.append(Spacer(1, 14 * mm))
    story.append(Paragraph("大模型测评报告", st_title))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(f"{name} · LLM Eval Bench v2 · evalscope 1.8.1", st_sub))
    story.append(Spacer(1, 10 * mm))

    # ===== 基本信息表 =====
    meta = [
        ["任务名称", name, "模型", config.get("model", "-")],
        ["接口格式", config.get("api_format", "-"), "状态", _status_cn(detail.get("status"))],
        ["运行时长", _dur(detail.get("duration")), "生成时间", _fmt_time(time.time())],
    ]
    mt = Table(meta, colWidths=[22 * mm, 58 * mm, 22 * mm, 60 * mm])
    mt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), CN), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), DIM), ("TEXTCOLOR", (2, 0), (2, -1), DIM),
        ("TEXTCOLOR", (1, 0), (1, -1), INK), ("TEXTCOLOR", (3, 0), (3, -1), INK),
        ("BACKGROUND", (0, 0), (0, -1), PRIMARY_LT),
        ("BACKGROUND", (2, 0), (2, -1), PRIMARY_LT),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, LINE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(mt)
    story.append(Spacer(1, 4 * mm))

    # 配置摘要
    cfg_lines = []
    ds_list = config.get("accuracy_datasets", []) or []
    if ds_list:
        cfg_lines.append(f"• 精度数据集（{len(ds_list)}）：{', '.join(ds_list[:8])}{'...' if len(ds_list) > 8 else ''}")
    cfg_lines.append(f"• Few-shot: {_n(config.get('few_shot', 0))}  |  抽样: {_s(config.get('sample_limit') or '全量')}  |  max_tokens: {_s(config.get('acc_max_tokens') or '自动')}  |  temperature: {config.get('acc_temperature', 0)}")
    if config.get("run_performance"):
        pc = config.get("perf", {}) or {}
        cfg_lines.append(f"• 性能压测: 并发 [{', '.join(str(x) for x in (pc.get('levels') or []))}] × {pc.get('requests_per_level', '-')} 请求  |  max_tokens: {pc.get('max_tokens', '-')}  |  {'流式' if pc.get('stream') else '非流式'}")
    if config.get("context_lengths"):
        cfg_lines.append(f"• 上下文扫描: {', '.join(str(x) for x in config['context_lengths'])} tokens  |  固定并发 {config.get('context_concurrency', '-')}")
    for line in cfg_lines:
        story.append(Paragraph(line, st_small))
    story.append(Spacer(1, 5 * mm))

    # ===== KPI 卡片 =====
    kpis = _build_kpis(result)
    if kpis:
        story.append(Paragraph("执行摘要", st_h2))
        kpi_count = len(kpis)
        col_w = (174 / kpi_count) * mm
        cells_top = [Paragraph(str(v), st_kpi) for v, _ in kpis]
        cells_bot = [Paragraph(l, st_kpi_lbl) for _, l in kpis]
        kt = Table([cells_top, cells_bot], colWidths=[col_w] * kpi_count)
        kt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), PRIMARY_LT),
            ("BOX", (0, 0), (-1, -1), 0.5, LINE),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, WHITE),
            ("TOPPADDING", (0, 0), (-1, 0), 10), ("BOTTOMPADDING", (0, 0), (-1, 0), 1),
            ("TOPPADDING", (0, 1), (-1, 1), 1), ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(kt)
        story.append(Spacer(1, 3 * mm))

    def section_table(data, col_widths, highlight_row=None):
        t = Table(data, colWidths=col_widths, hAlign="LEFT", repeatRows=1)
        ts = [
            ("FONTNAME", (0, 0), (-1, -1), CN), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, PRIMARY_LT]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, LINE),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"), ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (0, -1), 7),
        ]
        if highlight_row is not None:
            ts.append(("BACKGROUND", (0, highlight_row), (-1, highlight_row),
                      colors.HexColor("#D1FAE5")))
        t.setStyle(TableStyle(ts))
        return t

    # ===== 精度评测 =====
    if acc:
        story.append(Paragraph("一、精度评测结果", st_h2))
        data = [["数据集", "准确率", "题目数", "学科", "Few-shot"]]
        for ds, d in acc.items():
            shot = _n(d.get("few_shot", 0))
            acc_str = f"{d['accuracy']}%" if d.get("accuracy") is not None else "见详情"
            data.append([ds, acc_str,
                        str(_n(d.get("num") or d.get("total"))),
                        str(len(d.get("by_subject", {}) or {})),
                        f"{shot}-shot" if shot else "0-shot"])
        story.append(section_table(data, [48 * mm, 28 * mm, 28 * mm, 28 * mm, 42 * mm]))
        story.append(Spacer(1, 3 * mm))

        # 分学科/类别表
        for ds, d in acc.items():
            cats = d.get("by_category", {}) or {}
            subjs = d.get("by_subject", {}) or {}
            if cats:
                has_detail = any((cv.get("subsets") and len(cv["subsets"]) > 1) for cv in cats.values())
                sdata = [["类别", "准确率", "题目数", "包含学科"]]
                for cn, cv in cats.items():
                    sub_names = ", ".join(list((cv.get("subsets") or {}).keys())[:4])
                    if len(cv.get("subsets") or {}) > 4:
                        sub_names += "..."
                    sdata.append([
                        cn,
                        f"{cv['score']}%" if cv.get("score") is not None else "-",
                        str(_n(cv.get("num"))),
                        sub_names or "-",
                    ])
                if len(sdata) > 1:
                    story.append(Paragraph(f"{ds} · 类别得分", st_h3))
                    story.append(section_table(sdata, [44 * mm, 28 * mm, 24 * mm, 78 * mm]))
                    story.append(Spacer(1, 2 * mm))
            elif subjs and len(subjs) > 1:
                sdata = [["学科", "准确率"]]
                for sn, sv in sorted(subjs.items(), key=lambda x: -x[1]):
                    sdata.append([sn, f"{sv}%"])
                story.append(Paragraph(f"{ds} · 分学科准确率", st_h3))
                story.append(section_table(sdata, [120 * mm, 54 * mm]))
                story.append(Spacer(1, 2 * mm))

    # ===== 性能压测 =====
    if sweep:
        elems = [Paragraph("二、性能压测结果", st_h2)]
        best = perf.get("best", {}) or {}
        rec = perf.get("recommend")
        if best:
            parts = [f"吞吐峰值：并发 <b>{best.get('concurrency')}</b> 时达 <b>{best.get('rps')} RPS</b>（{_n(best.get('output_tps'))} tok/s）"]
            if rec:
                parts.append(f"推荐生产并发区间 <b>{rec['min']}~{rec['max']}</b>")
            elems.append(Paragraph(" · ".join(parts), st_body))
            elems.append(Spacer(1, 2 * mm))
        data = [["并发", "RPS", "输出tok/s", "TPOT", "TTFT", "平均延迟", "P90", "P99"]]
        best_idx = None
        for i, r in enumerate(sweep, 1):
            if best and r.get("concurrency") == best.get("concurrency"):
                best_idx = i
            data.append([
                str(r.get("concurrency")),
                str(_n(r.get("rps"))),
                str(_n(r.get("output_tps"))),
                f"{_n(r.get('tpot_avg_ms'))}ms" if r.get("tpot_avg_ms") else "-",
                f"{_n(r.get('ttft_avg'))}s" if r.get("ttft_avg") else "-",
                f"{_n(r.get('latency_avg'))}s",
                f"{_n(r.get('latency_p90'))}s",
                f"{_n(r.get('latency_p99'))}s",
            ])
        elems.append(section_table(data,
            [18*mm, 24*mm, 26*mm, 22*mm, 22*mm, 22*mm, 22*mm, 22*mm],
            highlight_row=best_idx))
        story.append(KeepTogether(elems))

        # 警告
        for w in (perf.get("warnings") or [])[:2]:
            story.append(Paragraph(
                f"<font color='#D97706'>⚠ {w.get('title', '')}</font>：{w.get('message', '')}",
                st_small))
        story.append(Spacer(1, 2 * mm))

    # ===== 上下文扫描 =====
    if ctx_sweep:
        story.append(Paragraph("三、上下文长度扫描结果", st_h2))
        story.append(Paragraph(
            f"固定并发 <b>{ctx.get('concurrency', '-')}</b>，每档 <b>{ctx.get('requests_per_level', '-')}</b> 请求",
            st_body))
        story.append(Spacer(1, 2 * mm))
        ctx_data = [["长度(tok)", "RPS", "输出tok/s", "TTFT", "平均延迟", "P99", "成功率"]]
        for r in ctx_sweep:
            ctx_len = r.get("context_length", 0)
            ctx_data.append([
                f"{ctx_len//1024}K" if ctx_len >= 1024 else str(ctx_len),
                str(_n(r.get("rps"))),
                str(_n(r.get("output_tps"))),
                f"{_n(r.get('ttft_avg'))}s" if r.get("ttft_avg") else "-",
                f"{_n(r.get('latency_avg'))}s",
                f"{_n(r.get('latency_p99'))}s",
                f"{_n(r.get('success_rate'))}%" if r.get("success_rate") else "-",
            ])
        story.append(section_table(ctx_data, [22*mm, 24*mm, 26*mm, 22*mm, 24*mm, 24*mm, 22*mm]))
        story.append(Spacer(1, 3 * mm))

    # ===== AI 分析 =====
    any_analysis = any(d.get("analysis", "").strip() for d in acc.values())
    if any_analysis:
        story.append(Paragraph("四、AI 分析报告", st_h2))
        for ds, d in acc.items():
            analysis = (d.get("analysis") or "").strip()
            if analysis:
                story.append(Paragraph(f"{ds}", st_h3))
                # 限制长度避免 PDF 过长
                lines = analysis.split("\n")
                # 取前 80 行，超出部分截断
                trimmed = "\n".join(lines[:80])
                if len(lines) > 80:
                    trimmed += "\n\n... (完整分析见 web 端)"
                for para in trimmed.split("\n"):
                    para = para.strip()
                    if para:
                        story.append(Paragraph(para, st_analysis))
                story.append(Spacer(1, 2 * mm))

    # ===== 推理性能统计 =====
    any_pm = any(isinstance(d.get("perf_metrics"), dict) for d in acc.values())
    if any_pm:
        story.append(Paragraph("五、推理性能统计", st_h2))
        for ds, d in acc.items():
            pm = d.get("perf_metrics")
            if not isinstance(pm, dict):
                continue
            lat = pm.get("latency", {}) or {}
            tp = pm.get("throughput", {}) or {}
            usage = pm.get("usage", {}) or {}
            in_tok = usage.get("input_tokens", {}) or {}
            out_tok = usage.get("output_tokens", {}) or {}

            story.append(Paragraph(
                f"{ds} · 样本 {_n(pm.get('n_samples'))} · "
                f"输出吞吐 {_n(tp.get('avg_output_tps'))} tok/s · "
                f"请求速率 {_n(tp.get('avg_req_ps'))} req/s",
                st_small))
            pm_data = [
                ["", "P25", "P50", "P75", "P90", "P99", "均值"],
                ["延迟(s)", _n(lat.get("25%")), _n(lat.get("50%")), _n(lat.get("75%")),
                 _n(lat.get("90%")), _n(lat.get("99%")), _n(lat.get("mean"))],
                ["输入Token", _n(in_tok.get("25%")), _n(in_tok.get("50%")), _n(in_tok.get("75%")),
                 _n(in_tok.get("90%")), _n(in_tok.get("99%")), _n(in_tok.get("mean"))],
                ["输出Token", _n(out_tok.get("25%")), _n(out_tok.get("50%")), _n(out_tok.get("75%")),
                 _n(out_tok.get("90%")), _n(out_tok.get("99%")), _n(out_tok.get("mean"))],
            ]
            story.append(section_table(pm_data,
                [24*mm, 22*mm, 22*mm, 22*mm, 22*mm, 22*mm, 22*mm]))
            story.append(Spacer(1, 2 * mm))

    if not acc and not sweep and not ctx_sweep:
        story.append(Paragraph("本任务暂无结果数据。", st_body))

    # ===== 附注 =====
    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=LINE, spaceAfter=5))
    story.append(Paragraph(
        "说明：本报告由大模型测评台自动生成。精度评测由业界标准工具 evalscope 执行，"
        "判分方式与各数据集官方一致，分数可对标公开榜单（请确认 few-shot 设置与榜单一致）。"
        "性能指标中 TPOT 为生成每 token 耗时、TTFT 为首字延迟、RPS 为每秒完成请求数。",
        st_small))

    doc.build(story, onFirstPage=on_first_page, onLaterPages=on_page)
    return path


# ============ 辅助函数 ============

def _status_cn(s):
    return {"done": "已完成", "running": "运行中", "error": "出错",
            "stopped": "已停止", "pending": "等待中"}.get(s, s or "-")


def _build_kpis(result: dict) -> list[tuple]:
    acc = result.get("accuracy", {})
    perf = result.get("performance", {})
    ctx = result.get("context_scan", {})
    kpis = []
    if acc:
        accs = [d.get("accuracy") for d in acc.values() if d.get("accuracy") is not None]
        if accs:
            avg = sum(accs) / len(accs)
            kpis.append((f"{avg:.1f}%", "平均准确率"))
        n_subjs = sum(len(d.get("by_subject", {}) or {}) for d in acc.values())
        if n_subjs:
            kpis.append((str(n_subjs), "学科覆盖"))
    if perf:
        best = perf.get("best") or {}
        if best:
            kpis.append((str(best.get("concurrency", "-")), "最优并发"))
            kpis.append((str(_n(best.get("rps"))), "峰值 RPS"))
    if ctx and ctx.get("sweep"):
        kpis.append((str(len(ctx["sweep"])), "上下文档位"))
    if acc:
        n_total = sum(_n(d.get("num") or d.get("total")) for d in acc.values())
        if n_total:
            kpis.append((str(n_total), "评测题数"))
    return kpis[:6]


def _fmt_subjects(subjs: dict) -> str:
    if not subjs:
        return "-"
    parts = []
    for ds, subs in subjs.items():
        if subs:
            parts.append(f"{ds}: {', '.join(subs[:4])}")
    return "; ".join(parts) if parts else "-"


def _fmt_perf_config(config: dict) -> str:
    if not config.get("run_performance"):
        return "未启用"
    pc = config.get("perf", {}) or {}
    parts = [
        f"并发 [{', '.join(str(x) for x in (pc.get('levels') or []))}]",
        f"×{pc.get('requests_per_level', '-')} 请求",
        f"max_tokens {pc.get('max_tokens', '-')}",
        '流式' if pc.get('stream') else '非流式',
    ]
    if pc.get("context_length"):
        parts.append(f"上下文 {pc['context_length']} tok")
    return "  |  ".join(parts)


def _fmt_ctx_config(config: dict) -> str:
    lengths = config.get("context_lengths") or []
    if not lengths:
        return "未启用"
    return f"{', '.join(str(x) for x in lengths)} tokens  |  并发 {config.get('context_concurrency', '-')}"